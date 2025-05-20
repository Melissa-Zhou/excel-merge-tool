import pandas as pd
import os
import glob
import tkinter as tk
from tkinter import filedialog, messagebox
from tkinter import ttk
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime, timedelta
import threading

# ---------- GUI 初始化 ----------
root = tk.Tk()
root.title("Excel 合并 + 抽取工具")
root.geometry("400x200")

label = tk.Label(root, text="选择文件夹合并 Excel 文件", wraplength=340)
label.pack(pady=10)

status_label = tk.Label(root, text="", fg="blue")
status_label.pack()

progress = ttk.Progressbar(root, orient="horizontal", length=300, mode="determinate")
progress.pack(pady=5)

def update_status(msg):
    status_label.config(text=msg)
    root.update_idletasks()

def merge_excel_and_prompt():
    threading.Thread(target=merge_task, daemon=True).start()

def merge_task():
    folder = filedialog.askdirectory(title="请选择包含 Excel 文件的文件夹")
    if not folder:
        return

    excel_files = glob.glob(os.path.join(folder, "*.xlsx"))
    total_files = len(excel_files)
    if total_files == 0:
        root.after(0, lambda: messagebox.showerror("错误", "未找到任何 .xlsx 文件！"))
        return

    merged_df = pd.DataFrame()
    header_template_file = None

    progress["maximum"] = total_files
    progress["value"] = 0

    for i, file in enumerate(excel_files):
        update_status(f"读取文件：{os.path.basename(file)} ({i+1}/{total_files})")
        try:
            df = pd.read_excel(file, sheet_name=0, header=0)
            merged_df = pd.concat([merged_df, df], ignore_index=True)
            if i == 0:
                header_template_file = file
        except Exception as e:
            root.after(0, lambda: messagebox.showerror("读取错误", f"读取文件 {file} 时出错：\n{e}"))
            return
        progress["value"] = i + 1
        update_status(f"已处理 {i+1}/{total_files} 个文件")
        root.update_idletasks()

    update_status("合并完成，正在保存格式化文件...")

    wb = load_workbook(header_template_file)
    ws_template = wb.active
    ws_template.delete_rows(2, ws_template.max_row - 1)
    total_rows = merged_df.shape[0]
    for r_idx, row in enumerate(dataframe_to_rows(merged_df, index=False, header=True), start=1):
        for c_idx, value in enumerate(row, start=1):
            ws_template.cell(row=r_idx, column=c_idx, value=value)
        if r_idx % 500 == 0 or r_idx == total_rows:
            update_status(f"正在写入第 {r_idx} / {total_rows} 行...")
            root.update_idletasks()

    merged_path = os.path.join(folder, "PPT_Report_Combine.xlsx")
    try:
        wb.save(merged_path)
        update_status("合并完成，文件已保存。")
        root.after(0, lambda: messagebox.showinfo("合并完成", f"文件保存为：\n{merged_path}"))
    except Exception as e:
        root.after(0, lambda: messagebox.showerror("保存失败", f"保存文件时出错：\n{e}"))
        return

    root.after(0, lambda: ask_filter_now(merged_path))

def ask_filter_now(merged_path):
    if messagebox.askyesno("是否继续？", "是否现在对合并结果进行筛选和抽取？"):
        threading.Thread(target=lambda: filter_and_extract(merged_path), daemon=True).start()

def filter_and_extract(file_path):
    update_status("正在读取合并结果文件...")
    progress["maximum"] = 4
    progress["value"] = 0
    root.update_idletasks()

    try:
        df = pd.read_excel(file_path, engine='openpyxl')
        progress["value"] += 1
        update_status("读取完成，准备筛选记录...")
        root.update_idletasks()
    except Exception as e:
        root.after(0, lambda: messagebox.showerror("读取错误", f"无法读取合并结果文件：\n{e}"))
        return

    date_col = "3000:二次确认(AD)"
    check_col = "7600:SAIS首检(AD)"

    today = datetime.today().date()
    date_7_days_ago = today - timedelta(days=7)
    df[date_col] = pd.to_datetime(df[date_col], errors='coerce')

    try:
        filtered = df[
            (df[date_col].dt.date >= date_7_days_ago) &
            (df[date_col].dt.date < today) &
            (df[check_col].isna())
        ]
        progress["value"] += 1
        update_status(f"筛选完成：共找到 {len(filtered)} 条记录")
        root.update_idletasks()
    except Exception as e:
        root.after(0, lambda: messagebox.showerror("筛选失败", f"筛选出错：\n{e}"))
        return

    if filtered.empty:
        update_status("筛选结果为空")
        root.after(0, lambda: messagebox.showinfo("结果为空", "没有符合筛选条件的记录。"))
        return

    try:
        update_status("正在抽取 20% 样本...")
        sample_df = filtered.sample(frac=0.2)
        progress["value"] += 1
        root.update_idletasks()
    except Exception as e:
        root.after(0, lambda: messagebox.showerror("抽取失败", f"抽样出错：\n{e}"))
        return

    selected_columns = ["销售凭证", "安装区域名称", "安装分公司名称", "Eq.PE", "Eq.PE - 名称"]
    try:
        result = sample_df[selected_columns]
    except KeyError as e:
        root.after(0, lambda: messagebox.showerror("列缺失", f"结果中找不到以下列：\n{e}"))
        return

    output_path = os.path.join(os.path.dirname(file_path), "PE_ICL抽取结果.xlsx")
    try:
        update_status("正在保存抽取结果...")
        result.to_excel(output_path, index=False)
        progress["value"] += 1
        update_status("筛选+抽取流程完成。")
        root.after(0, lambda: messagebox.showinfo("筛选完成", f"结果保存为：\n{output_path}"))
    except Exception as e:
        root.after(0, lambda: messagebox.showerror("保存失败", f"保存文件时出错：\n{e}"))

# 按钮绑定
button = tk.Button(root, text="开始合并", command=merge_excel_and_prompt)
button.pack(pady=10)

root.mainloop()
